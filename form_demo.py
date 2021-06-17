# -*- coding = utf-8 -*-
# @Time : 2021/6/17 23:51
# @Author : toast
# @File : form_demo.py
# @Software : PyCharm

import openpyxl

# 1. 打开一个excel文档, class 'openpyxl.workbook.workbook.Workbook'实例化出来的对象
wb = openpyxl.load_workbook('test1.xlsx')
print(wb, type(wb))

# 获取当前工作薄里所有的工作表,和正在使用的表;
print("当前工作簿中所有的工作表为：%s" % wb.sheetnames)
print("正在使用的工作表为：%s" % wb.active)

# 2. 选择要操作的工作表， 返回工作表对象
sheet = wb['Sheet1']
# 获取工作表的名称
print("工作表的名称为：%s" % sheet.title)

# 3. 返回指定行指定列的单元格信息
print("第1行第2列的单元信息为：%s" % sheet.cell(row=1, column=2).value)

cell = sheet['B1']
print("cell对应的信息为 %s" % cell)
print("cell的行为：%s,列为：%s, 单元格信息为：%s" % (cell.row, cell.column, cell.value))

# 4. 获取工作表中行和列的最大值
print("工作表有%s行" % sheet.max_column)
print("工作表有%s列" % sheet.max_row)
# 更改工作表的名称
sheet.title = 'Sheet1'
print(sheet.title)

# 5. 访问单元格的所有信息
print(sheet.rows)  # 返回一个生成器， 包含文件的每一行内容， 可以通过便利访问.
# 循环遍历每一行
for row in sheet.rows:
    # 循环遍历每一个单元格
    for cell in row:
        # 获取单元格的内容
        print(cell.value, end=',')
    print()

# 6. 保存修改信息
wb.save(filename='test1.xlsx')
print("信息已近保存。")

# 操作Excel表格可详细的概括如下：
# 1.导入 openpyxl 模块。
# 2.调用 openpyxl.load_workbook()函数。
# 3.取得 Workbook 对象。
# 4.调用 wb.sheetnames和 wb.active 获取工作簿详细信息。
# 5.取得 Worksheet 对象。
# 6.使用索引或工作表的 cell()方法,带上 row 和 column 关键字参数。
# 7.取得 Cell 对象。
# 8.读取 Cell 对象的 value 属性

